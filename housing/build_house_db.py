import sys, os
import openpyxl as pyxl
import datetime
import pandas as pd

filename = '5HundredFalls_VXII_020216.xlsm'

def parse_excel(filename, 
		filedir = '/home/matt/singularity/housing/r_housing/'):
	"""
	Function: 	parse_excel(ws)
	Input:		filename: an excel document
				filedir: location of the file (location
					in Matt's file system by default) 
	Output:		appends the data from the worksheets
					to an appropriate .csv file
	"""
	full_file = filedir + filename
	print 'Reading: ' + full_file
	wb = pyxl.load_workbook(full_file, read_only = True,
		data_only = True)
	print  'Reading complete ... '
	worksheets = wb.worksheets
	ws = wb.get_sheet_by_name('Data_Input')
	for worksheet in worksheets:

		if worksheet.title == 'Data_Input':
			print_out =  'Parsing file: ' + filename
			print_out += ' sheet: ' + worksheet.title
			print print_out
			data_dict = {
				'property_name' : ws['C3'].value,
				'property_street' : ws['C4'].value,
				'property_city' : ws['C5'].value,
				'total_units' : ws['C6'].value,
				'occupied_units' : ws['F59'].value,
				'vacant_units' : ws['G59'].value,
				'total_sq_ft' : ws['C7'].value,
				'avg_sq_ft' : ws['F3'].value,
				'completion_date' : ws['F5'].value,
				'num_buildings' : ws['F6'].value,
				'attributes' : ws['F7'].value,
				'land_area' : ws['J3'].value,
				'density' : ws['J4'].value,
				'analysis_begins' : ws['M3'].value,
				'first_fy_ends' : ws['M4'].value,
				'hold_period' : ws['M5'].value,
				'proposed_rent_effective' : ws['M19'].value,
				'rent_per_unit' : ws['H59'].value,
				'proposed_rent' : ws['I59'].value
			}
			print 'Parsing complete ... '
	return data_dict	

if __name__ == '__main__':
	filedir = '/home/matt/singularity/housing/r_housing'
	os.chdir(filedir)
	# Find the names of the tabs that are in
	#	the UW template file
	#file_name = file_dir + '/UWTemplate_VXII_030916.xlsm'
	#wb = pyxl.load_workbook(file_name)
	#temp_worksheets = wb.worksheets
	#template_titles = [x.title for x in temp_worksheets]
	parsed = []
	not_parsed = []
	dict_list = []
	for file in os.listdir(filedir):
		if file != 'UWTemplate_VXII_030916.xlsm':
			try:
				data_dict = parse_excel(file)
				dict_list.append(data_dict)
				parsed.append(file)
			except:
				print 'Could not parse ' + file
				not_parsed.append(file)
				continue
	print 'Successfully parsed ' + str(len(parsed)) + ' files. '
	print 'Could not parse ' + str(len(not_parsed)) + ' files.'
	agg_dict = {}
	for key in dict_list[0]:
		agg_dict[key] = [d[key] for d in dict_list]
	df = pd.DataFrame(agg_dict)
	parsedir = '/home/matt/singularity/housing/r_housing_parsed'
	os.chdir(parsedir)
	df.to_csv('r_data_input.csv')

